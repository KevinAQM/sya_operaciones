# sya_operaciones_server.py
import os
import logging
from datetime import datetime
import pandas as pd
import openpyxl
from flask import Flask, request, jsonify, send_file

app = Flask(__name__)

# Usar una ruta absoluta para el archivo Excel
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "registros_trabajo.xlsx")
MATERIALES_CSV_PATH = os.path.join(BASE_DIR, "operaciones_materiales.csv")
EQUIPOS_CSV_PATH = os.path.join(BASE_DIR, "operaciones_equipos.csv")
VEHICULOS_CSV_PATH = os.path.join(BASE_DIR, "operaciones_vehiculos.csv")
PERSONAL_CSV_PATH = os.path.join(BASE_DIR, "operaciones_personal.csv")

# Configuración de logging
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

def inicializar_excel():
    if not os.path.exists(EXCEL_FILE):
        logging.info(f"Creando archivo Excel en: {EXCEL_FILE}")
        wb = openpyxl.Workbook()
        # Hoja "Reporte Principal"
        ws1 = wb.active
        ws1.title = "Reporte Principal"
        headers_reporte = [
            "Fecha", "Código Obra", "Nombre Ingeniero",
            "Nombre Supervisor", "Actividad Principal",
            "Supervisor Presente", "Equipos Seguridad Completos",
            "Incidentes", "Plan Siguiente Día", "Observaciones"
        ]
        ws1.append(headers_reporte)

        # Hoja "Materiales Usados"
        ws2 = wb.create_sheet(title="Materiales Usados")
        headers_materiales = [
            "Fecha", "Código Obra", "Nombre Ingeniero"
        ]
        ws2.append(headers_materiales)

        # Hoja "Equipos Usados"
        ws3 = wb.create_sheet(title="Equipos Usados")
        headers_equipos = [
            "Fecha", "Código Obra", "Nombre Ingeniero"
        ]
        ws3.append(headers_equipos)

        # Hoja "Vehículos Usados"
        ws4 = wb.create_sheet(title="Vehículos Usados")
        headers_vehiculos = [
            "Fecha", "Código Obra", "Nombre Ingeniero"
        ]
        ws4.append(headers_vehiculos)

        # Hoja "Personal de Campo"
        ws5 = wb.create_sheet(title="Personal de Campo")
        headers_personal = [
            "Fecha", "Código Obra", "Nombre Ingeniero"
        ]
        ws5.append(headers_personal)

        wb.save(EXCEL_FILE)
    else:
        logging.info(f"El archivo Excel ya existe en: {EXCEL_FILE}")

def actualizar_cabeceras_materiales(ws, num_materiales):
    headers = ws[1]
    num_headers_actuales = len(headers)

    ultimo_material = 0
    for header in headers:
        header_value = header.value
        if header_value and header_value.startswith("Material"):
            try:
                numero = int(header_value.split(" ")[1])
                ultimo_material = max(ultimo_material, numero)
            except ValueError:
                pass

    if ultimo_material >= num_materiales:
        return

    for i in range(ultimo_material + 1, num_materiales + 1):
        ws.cell(row=1, column=num_headers_actuales + 1, value=f"Material {i}")
        ws.cell(row=1, column=num_headers_actuales + 2, value=f"Unidad {i}")
        ws.cell(row=1, column=num_headers_actuales + 3, value=f"Cantidad {i}")
        num_headers_actuales += 3

def actualizar_cabeceras_equipos(ws, num_equipos):
    headers = ws[1]
    num_headers_actuales = len(headers)

    ultimo_equipo = 0
    for header in headers:
        header_value = header.value
        if header_value and header_value.startswith("Equipo"):
            try:
                numero = int(header_value.split(" ")[1])
                ultimo_equipo = max(ultimo_equipo, numero)
            except ValueError:
                pass

    if ultimo_equipo >= num_equipos:
        return

    for i in range(ultimo_equipo + 1, num_equipos + 1):
        ws.cell(row=1, column=num_headers_actuales + 1, value=f"Equipo {i}")
        ws.cell(row=1, column=num_headers_actuales + 2, value=f"Cantidad {i}")
        ws.cell(row=1, column=num_headers_actuales + 3, value=f"Propiedad {i}")
        num_headers_actuales += 3

def actualizar_cabeceras_vehiculos(ws, num_vehiculos):
    headers = ws[1]
    num_headers_actuales = len(headers)

    ultimo_vehiculo = 0
    for header in headers:
        header_value = header.value
        if header_value and header_value.startswith("Vehículo"):
            try:
                numero = int(header_value.split(" ")[1])
                ultimo_vehiculo = max(ultimo_vehiculo, numero)
            except ValueError:
                pass

    if ultimo_vehiculo >= num_vehiculos:
        return

    for i in range(ultimo_vehiculo + 1, num_vehiculos + 1):
        ws.cell(row=1, column=num_headers_actuales + 1, value=f"Vehículo {i}")
        ws.cell(row=1, column=num_headers_actuales + 2, value=f"Placa {i}")
        ws.cell(row=1, column=num_headers_actuales + 3, value=f"Propiedad {i}")
        num_headers_actuales += 3

def actualizar_cabeceras_personal(ws, num_personal):
    headers = ws[1]
    num_headers_actuales = len(headers)

    ultimo_personal = 0
    for header in headers:
        header_value = header.value
        if header_value and header_value.startswith("Personal"):
            try:
                numero = int(header_value.split(" ")[1])
                ultimo_personal = max(ultimo_personal, numero)
            except ValueError:
                pass

    if ultimo_personal >= num_personal:
        return

    for i in range(ultimo_personal + 1, num_personal + 1):
        ws.cell(row=1, column=num_headers_actuales + 1, value=f"Personal {i}")
        ws.cell(row=1, column=num_headers_actuales + 2, value=f"Categoría {i}")
        ws.cell(row=1, column=num_headers_actuales + 3, value=f"Horas extras {i}")
        num_headers_actuales += 3


def procesar_datos(datos):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws_reporte = wb["Reporte Principal"]
        ws_materiales = wb["Materiales Usados"]
        ws_equipos = wb["Equipos Usados"]
        ws_vehiculos = wb["Vehículos Usados"]
        ws_personal = wb["Personal de Campo"]


        # Preparar fila de datos para "Reporte Principal"
        fila_reporte = [
            # Formatear la fecha como DATE en "Reporte Principal"
            datetime.strptime(datos.get('fecha', ''), '%d/%m/%Y').date(),  # Fecha como date
            datos.get('codigo_obra', ''),
            datos.get('nombre_ingeniero', ''),
            datos.get('nombre_supervisor', ''),
            datos.get('actividad_principal', ''),
            'Sí' if datos.get('supervisor_presente', False) else 'No',
            'Sí' if datos.get('equipos_seguridad', False) else 'No',
            datos.get('incidentes', ''),
            datos.get('siguiente_dia', ''),
            datos.get('observaciones', '')
        ]
        ws_reporte.append(fila_reporte)

        # Procesar materiales usados para "Materiales Usados"
        materiales = datos.get('materiales_usados', [])
        actualizar_cabeceras_materiales(ws_materiales, len(materiales))
        fila_materiales = [
            datetime.strptime(datos.get('fecha', ''), '%d/%m/%Y').date(),
            datos.get('codigo_obra', ''),
            datos.get('nombre_ingeniero', '')
        ]
        for material in materiales:
            fila_materiales.extend([material['nombre'], material['unidad'], material['cantidad']])
        ws_materiales.append(fila_materiales)

        # Procesar equipos usados
        equipos = datos.get('equipos_usados', [])
        actualizar_cabeceras_equipos(ws_equipos, len(equipos))
        fila_equipos = [
            datetime.strptime(datos.get('fecha', ''), '%d/%m/%Y').date(),
            datos.get('codigo_obra', ''),
            datos.get('nombre_ingeniero', '')
        ]
        for equipo in equipos:
            fila_equipos.extend([equipo['nombre'], equipo['cantidad'], equipo['propiedad']])
        ws_equipos.append(fila_equipos)

        # Procesar vehículos usados
        vehiculos = datos.get('vehiculos_usados', [])
        actualizar_cabeceras_vehiculos(ws_vehiculos, len(vehiculos))
        fila_vehiculos = [
            datetime.strptime(datos.get('fecha', ''), '%d/%m/%Y').date(),
            datos.get('codigo_obra', ''),
            datos.get('nombre_ingeniero', '')
        ]
        for vehiculo in vehiculos:
            fila_vehiculos.extend([vehiculo['nombre'], vehiculo['placa'], vehiculo['propiedad']])
        ws_vehiculos.append(fila_vehiculos)

        # Procesar personal de campo
        personal_campo = datos.get('personal_de_campo', [])
        actualizar_cabeceras_personal(ws_personal, len(personal_campo))
        fila_personal = [
            datetime.strptime(datos.get('fecha', ''), '%d/%m/%Y').date(),
            datos.get('codigo_obra', ''),
            datos.get('nombre_ingeniero', '')
        ]
        for personal in personal_campo:
            fila_personal.extend([personal['nombre_completo'], personal['categoria'], personal['horas_extras']])
        ws_personal.append(fila_personal)


        wb.save(EXCEL_FILE)
        logging.info(f"Datos recibidos de {datos.get('nombre_ingeniero', 'Unknown')} procesados exitosamente")

    except Exception as e:
        logging.error(f"Error al procesar datos: {str(e)}")

def descargar_excel_flask():
    try:
        logging.info(f"Intentando enviar archivo: {EXCEL_FILE}")
        return send_file(EXCEL_FILE, as_attachment=True)
    except Exception as e:
        logging.error(f"Error al generar descarga de Excel: {str(e)}")
        return str(e), 500

def agregar_nuevo_material_csv(nombre_material, unidad):
    try:
        df = pd.read_csv(MATERIALES_CSV_PATH)
        nuevo_material = pd.DataFrame([{'nombre_material': nombre_material, 'unidad': unidad}])
        df = pd.concat([df, nuevo_material], ignore_index=True)
        df.to_csv(MATERIALES_CSV_PATH, index=False)
        logging.info(f"Nuevo material '{nombre_material}' agregado a {MATERIALES_CSV_PATH}")
        return True
    except Exception as e:
        logging.error(f"Error al agregar nuevo material a CSV: {str(e)}")
        return False

def agregar_nuevo_equipo_csv(nombre_equipo, propiedad):
    try:
        df = pd.read_csv(EQUIPOS_CSV_PATH)
        nuevo_equipo = pd.DataFrame([{'nombre_equipo': nombre_equipo, 'propiedad': propiedad}])
        df = pd.concat([df, nuevo_equipo], ignore_index=True)
        df.to_csv(EQUIPOS_CSV_PATH, index=False)
        logging.info(f"Nuevo equipo '{nombre_equipo}' agregado a {EQUIPOS_CSV_PATH}")
        return True
    except Exception as e:
        logging.error(f"Error al agregar nuevo equipo a CSV: {str(e)}")
        return False

def agregar_nuevo_vehiculo_csv(nombre_vehiculo, placa, propiedad):
    try:
        df = pd.read_csv(VEHICULOS_CSV_PATH)
        nuevo_vehiculo = pd.DataFrame([{'nombre_vehiculo': nombre_vehiculo, 'placa': placa, 'propiedad': propiedad}])
        df = pd.concat([df, nuevo_vehiculo], ignore_index=True)
        df.to_csv(VEHICULOS_CSV_PATH, index=False)
        logging.info(f"Nuevo vehículo '{nombre_vehiculo}' agregado a {VEHICULOS_CSV_PATH}")
        return True
    except Exception as e:
        logging.error(f"Error al agregar nuevo vehículo a CSV: {str(e)}")
        return False

def agregar_nuevo_personal_csv(apellido_paterno, apellido_materno, nombres, categoria):
    try:
        df = pd.read_csv(PERSONAL_CSV_PATH)
        nuevo_personal = pd.DataFrame([{
            'AP. PATERNO': apellido_paterno,
            'AP. MATERNO': apellido_materno,
            'NOMBRES': nombres,
            'CATEGORIA': categoria
        }])
        df = pd.concat([df, nuevo_personal], ignore_index=True)
        df.to_csv(PERSONAL_CSV_PATH, index=False)
        logging.info(f"Nuevo personal '{nombres} {apellido_paterno}' agregado a {PERSONAL_CSV_PATH}")
        return True
    except Exception as e:
        logging.error(f"Error al agregar nuevo personal a CSV: {str(e)}")
        return False


# Inicializar Excel al inicio
inicializar_excel()

@app.route('/api/materiales', methods=['GET'])
def get_materiales():
    try:
        df = pd.read_csv(MATERIALES_CSV_PATH)
        materiales = df.to_dict(orient='records')
        return jsonify(materiales)
    except FileNotFoundError:
        return jsonify({"error": f"No se encontró el archivo de materiales en {MATERIALES_CSV_PATH}"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/equipos', methods=['GET'])
def get_equipos():
    try:
        df = pd.read_csv(EQUIPOS_CSV_PATH)
        equipos = df.to_dict(orient='records')
        return jsonify(equipos)
    except FileNotFoundError:
        return jsonify({"error": f"No se encontró el archivo de equipos en {EQUIPOS_CSV_PATH}"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/vehiculos', methods=['GET'])
def get_vehiculos():
    try:
        df = pd.read_csv(VEHICULOS_CSV_PATH)
        vehiculos = df.to_dict(orient='records')
        return jsonify(vehiculos)
    except FileNotFoundError:
        return jsonify({"error": f"No se encontró el archivo de vehículos en {VEHICULOS_CSV_PATH}"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/personal', methods=['GET'])
def get_personal():
    try:
        df = pd.read_csv(PERSONAL_CSV_PATH)
        personal = df.to_dict(orient='records')
        return jsonify(personal)
    except FileNotFoundError:
        return jsonify({"error": f"No se encontró el archivo de personal en {PERSONAL_CSV_PATH}"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/materiales/new', methods=['POST'])
def new_material():
    data = request.json
    if not data or 'nombre_material' not in data or 'unidad' not in data:
        return jsonify({"error": "Nombre de material y unidad son requeridos"}), 400
    if agregar_nuevo_material_csv(data['nombre_material'], data['unidad']):
        return jsonify({"status": "success"}), 201
    else:
        return jsonify({"error": "Error al agregar nuevo material"}), 500

@app.route('/api/equipos/new', methods=['POST'])
def new_equipo():
    data = request.json
    if not data or 'nombre_equipo' not in data or 'propiedad' not in data:
        return jsonify({"error": "Nombre de equipo y propiedad son requeridos"}), 400
    if agregar_nuevo_equipo_csv(data['nombre_equipo'], data['propiedad']):
        return jsonify({"status": "success"}), 201
    else:
        return jsonify({"error": "Error al agregar nuevo equipo"}), 500

@app.route('/api/vehiculos/new', methods=['POST'])
def new_vehiculo():
    data = request.json
    if not data or 'nombre_vehiculo' not in data or 'placa' not in data or 'propiedad' not in data:
        return jsonify({"error": "Nombre de vehículo, placa y propiedad son requeridos"}), 400
    if agregar_nuevo_vehiculo_csv(data['nombre_vehiculo'], data['placa'], data['propiedad']):
        return jsonify({"status": "success"}), 201
    else:
        return jsonify({"error": "Error al agregar nuevo vehículo"}), 500

@app.route('/api/personal/new', methods=['POST'])
def new_personal():
    data = request.json
    if not data or 'nombre_completo' not in data or 'categoria' not in data:
        return jsonify({"error": "Nombre completo y categoría de personal son requeridos"}), 400

    nombre_completo = data['nombre_completo']
    partes_nombre = nombre_completo.split(',')
    if len(partes_nombre) != 2:
        return jsonify({"error": "Formato de nombre incorrecto. Debe ser 'apellido_paterno apellido_materno, nombres'"}), 400

    apellidos_str, nombres = partes_nombre
    apellidos_partes = apellidos_str.strip().split()
    if not apellidos_partes:
        return jsonify({"error": "Apellido paterno requerido"}), 400
    apellido_paterno = apellidos_partes[0]
    apellido_materno = apellidos_partes[1] if len(apellidos_partes) > 1 else ''

    if agregar_nuevo_personal_csv(apellido_paterno, apellido_materno, nombres.strip(), data['categoria']):
        return jsonify({"status": "success"}), 201
    else:
        return jsonify({"error": "Error al agregar nuevo personal"}), 500


@app.route('/recibir-datos', methods=['POST'])
def recibir_datos():
    datos = request.json
    procesar_datos(datos)
    return jsonify({"status": "success"})

@app.route('/descargar-excel', methods=['GET'])
def descargar_excel_route():
    return descargar_excel_flask()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)